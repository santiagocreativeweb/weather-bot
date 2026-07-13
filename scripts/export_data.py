#!/usr/bin/env python3
# scripts/export_data.py — Base de datos SQLite + export a Excel de TODO el sistema (objetivos #7/#8).
# [Creado 2026-07-10, Bloque C.]
#
# QUE HACE (idempotente, corre a diario): consolida los CSV/JSON append-only del sistema (que son la
# FUENTE de verdad) en:
#   1) data/wxbt.db      — base SQLite consultable (una tabla por fuente + leaderboard + meta).
#   2) data/wxbt_export.xlsx — libro Excel con una hoja por dominio, formateado (header, autofiltro,
#      panel congelado, anchos), + una hoja Resumen con las metricas clave.
# Las tablas se REEMPLAZAN en cada corrida (los CSV son el histórico; el DB es una vista materializada).
# No toca ningún archivo existente; solo escribe wxbt.db y wxbt_export.xlsx.
#
# USO: python scripts/export_data.py            (usa la fecha real como sello)
#      python scripts/export_data.py --date YYYY-MM-DD
import argparse, json, os, sqlite3, sys
import datetime as dt
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
D = os.path.join(HERE, "..", "data")
DB = os.path.join(D, "wxbt.db")
XLSX = os.path.join(D, "wxbt_export.xlsx")
LOG = os.path.join(D, "accumulator.log")

# (archivo CSV, nombre de tabla SQLite, nombre de hoja Excel). El orden es el de las hojas del libro.
SOURCES = [
    ("predictions_forward.csv", "predictions", "Predicciones"),
    ("pred_scores.csv",         "pred_scores", "Resultados"),
    ("backfill_check.csv",      "backfill",    "Backfill (lab)"),
    ("books_forward.csv",       "books",       "Books"),
    ("ensemble_forward.csv",    "ensemble",    "Ensemble"),
    ("city_scout.csv",          "city_scout",  "Ciudades (scout)"),
    ("nbm_forward.csv",         "nbm",         "Fuente NBM"),
    ("mosmix_tx_forward.csv",   "mosmix_tx",   "Fuente MOSMIX"),
    ("cwa_forward.csv",         "cwa",         "Fuente CWA"),
]


def log_run(status, detail):
    ts = dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")
    with open(LOG, "a", newline="") as f:
        f.write(f"{ts} | export | {dt.date.today().isoformat()} | {status} | {detail}\n")


def read_csv(fname):
    path = os.path.join(D, fname)
    if not os.path.exists(path) or os.path.getsize(path) == 0:
        return None
    try:
        return pd.read_csv(path)
    except Exception as e:
        print(f"[WARN] {fname}: {e}", file=sys.stderr)
        return None


def load_bias():
    """station_bias.json -> DataFrame (station, bias_v2) + dict de metadatos."""
    path = os.path.join(D, "station_bias.json")
    if not os.path.exists(path):
        return None, {}
    try:
        j = json.load(open(path, encoding="utf-8"))
        df = pd.DataFrame(sorted(j.get("bias", {}).items()), columns=["station", "bias_v2"])
        meta = {k: j.get(k) for k in ("asof", "window_days", "variant", "note")}
        return df, meta
    except Exception as e:
        print(f"[WARN] station_bias.json: {e}", file=sys.stderr)
        return None, {}


def leaderboard_df(backfill):
    """Ranking de estaciones (MISMA formula que scripts/leaderboard.py: score = hit*100 − mae*8 −
    std*6, sobre backfill lead-2). Se recomputa aca para que el xlsx sea autocontenido."""
    if backfill is None or backfill.empty:
        return None
    bf = backfill[(backfill.lead == 2) & backfill.max_real.notna()].copy()
    if bf.empty:
        return None
    bf["ae"] = (bf.mu_cal - bf.max_real).abs()
    bf["e"] = bf.mu_cal - bf.max_real
    rows = []
    for st, g in bf.groupby("station"):
        hit = g.hit_cal.mean() if "hit_cal" in g else float("nan")
        rows.append(dict(station=st, n=len(g), hit_exacto=round(hit, 4),
                         mae=round(g.ae.mean(), 3), sesgo=round(g.e.mean(), 3),
                         estabilidad=round(g.e.std(), 3),
                         score=round(hit * 100 - g.ae.mean() * 8 - g.e.std() * 6, 2)))
    return pd.DataFrame(rows).sort_values("score", ascending=False).reset_index(drop=True)


def resumen_df(tables, stamp):
    """Hoja Resumen: fila por fuente con conteo + rango de fechas, encabezada por el sello."""
    rows = [dict(seccion="generado", detalle=stamp, valor="")]
    for name, df in tables.items():
        if df is None:
            rows.append(dict(seccion=name, detalle="sin datos", valor=0)); continue
        rango = ""
        for col in ("target", "snapshot_date", "avail_utc", "capture_utc", "sent_utc"):
            if col in df.columns and df[col].notna().any():
                rango = f"{df[col].min()} .. {df[col].max()}"; break
        rows.append(dict(seccion=name, detalle=rango, valor=len(df)))
    return pd.DataFrame(rows)


def write_xlsx(sheets, stamp):
    """sheets = lista [(nombre_hoja, df)]. Formato: header en negrita/fondo, panel congelado en la
    fila 1, autofiltro, ancho de columnas por contenido. Hoja vacia -> placeholder."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(XLSX, engine="openpyxl") as xw:
        for name, df in sheets:
            safe = name[:31]   # Excel: max 31 chars por hoja
            if df is None or df.empty:
                pd.DataFrame({"info": ["sin datos disponibles"]}).to_excel(xw, sheet_name=safe, index=False)
                continue
            df.to_excel(xw, sheet_name=safe, index=False)
            ws = xw.sheets[safe]
            hdr_fill = PatternFill("solid", fgColor="1F3A5F")
            hdr_font = Font(color="FFFFFF", bold=True, size=10)
            for c in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill = hdr_fill; cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"
            for i, col in enumerate(df.columns, 1):
                width = max(len(str(col)), *(df[col].astype(str).str.len().fillna(0).tolist() or [0]))
                ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 8), 42)


def main(a):
    stamp = (a.date or dt.date.today().isoformat()) + " " + \
        dt.datetime.now(dt.timezone.utc).strftime("%H:%M UTC")
    tables = {}
    con = sqlite3.connect(DB)
    try:
        for fname, tbl, _sheet in SOURCES:
            df = read_csv(fname)
            tables[tbl] = df
            if df is not None:
                df.to_sql(tbl, con, if_exists="replace", index=False)
        # calibraciones (bias V2) + leaderboard computado
        bias_df, bias_meta = load_bias()
        if bias_df is not None:
            bias_df.to_sql("station_bias", con, if_exists="replace", index=False)
        lb = leaderboard_df(tables.get("backfill"))
        if lb is not None:
            lb.to_sql("leaderboard", con, if_exists="replace", index=False)
        # meta
        meta_rows = [("generado", stamp)] + [(f"n_{t}", str(len(df) if df is not None else 0))
                                             for t, df in tables.items()]
        meta_rows += [(f"bias_{k}", str(v)) for k, v in bias_meta.items()]
        pd.DataFrame(meta_rows, columns=["clave", "valor"]).to_sql("meta", con, if_exists="replace", index=False)
        con.commit()
    finally:
        con.close()

    # Excel: Resumen + Leaderboard primero, luego las fuentes, + Calibraciones al final
    sheets = [("Resumen", resumen_df(tables, stamp)),
              ("Leaderboard", lb)]
    for fname, tbl, sheet in SOURCES:
        sheets.append((sheet, tables.get(tbl)))
    if bias_df is not None:
        cal = bias_df.copy()
        for k, v in bias_meta.items():
            cal[k] = v   # anexar metadatos del bias (asof/variant/note) como columnas de contexto
        sheets.append(("Calibraciones", cal))
    write_xlsx(sheets, stamp)

    n_tabs = sum(1 for _, df in tables.items() if df is not None) + (1 if bias_df is not None else 0)
    print(f"SQLite -> {os.path.abspath(DB)} ({n_tabs} tablas de datos + leaderboard + meta)")
    print(f"Excel  -> {os.path.abspath(XLSX)} ({len(sheets)} hojas)")
    log_run("OK", f"tablas={n_tabs} hojas={len(sheets)}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Consolida los CSV del sistema en SQLite + Excel (#7/#8).")
    ap.add_argument("--date", default=None, help="sello de fecha YYYY-MM-DD (default: hoy)")
    main(ap.parse_args())
